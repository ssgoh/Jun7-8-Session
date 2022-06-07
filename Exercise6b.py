#Exercise6b.py
#Library
from time import sleep
import board
import adafruit_dht
#pip install openpyxl  - to install excel library
from openpyxl import Workbook, load_workbook
from os.path import exists
import datetime

#Setup
dhtDevice=adafruit_dht.DHT22(board.D17,use_pulseio=False)
excel_filename='weatherrecord.xlsx'

#Function
def getReadingDateAndTime():
    today=datetime.datetime.now()
    read_date= today.strftime("%d") +'-'+today.strftime("%B") + "-" + today.strftime("%Y")
    read_time= today.strftime("%H") +':'+today.strftime("%M") + ":" + today.strftime("%S") 
    return read_date,read_time

def readDHT22():
    try:
        #sensor reads correctly
        temperature_c = dhtDevice.temperature
        humidity=dhtDevice.humidity
    except:
        #sensor mulfunction
        temperature_c=0
        humidity=0
    return (humidity,temperature_c)

#Algorithm
while True:
    humidity,temperature=readDHT22()
    print(humidity,temperature)
    day,time = getReadingDateAndTime()
    
    #Save Data to Excel
    if exists(excel_filename):
        wb = load_workbook(excel_filename)
        page=wb.active
    else:
        #Create a new workbook and Headers for the Columns
        wb = Workbook()
        page = wb.active
        headers = ['Date','Time','Humidity','Temperature']
        page.append(headers)
    
    newdata=[]
    newdata.append(day)
    newdata.append(time)
    newdata.append(humidity)
    newdata.append(temperature)
    page.append(newdata)
    wb.save(filename=excel_filename)
    sleep(5)